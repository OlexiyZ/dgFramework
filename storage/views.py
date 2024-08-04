from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpRequest, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from openpyxl import load_workbook
import json
import pandas as pd
import psycopg2
from django.core.files.storage import FileSystemStorage
import os
from .models import *

wb = None


# Create your views here.
# @csrf_exempt
def excelImport(request: HttpRequest):
    context = {
        "text": "Excel Import!!!"
    }
    return render(request, 'storage/excelimport.html', context)


@csrf_exempt
# @require_POST
def upload_file(request):
    global wb
    # context = {'message': 'Файл успешно загружен'}
    context = {}
    if request.FILES.get('excelFile'):
        file = request.FILES.get('excelFile')
        if not file:
            return JsonResponse({'error': 'No file to download'}, status=400)

        # Здесь вы можете обрабатывать файл, например, сохранять его на сервере
        fss = FileSystemStorage()
        filename = fss.save(file.name, file)
        # file_url = fss.url(filename)
        file_path = fss.path(filename)
        context = {'message': 'File uploaded successfully', 'file_path': file_path}
        wb = load_workbook(file)

        sheets = dict()
        # Adding items to Listbox
        for sheet_item in wb.worksheets:
            tables = []
            sheet_title = sheet_item.title
            # sheets.append(sheet_title)
            selected_sheet = wb[sheet_title]
            for table_item in selected_sheet.tables:
                tables.append(table_item)
            sheets[sheet_title] = tables

        context['sheets'] = sheets
        # return JsonResponse(sheets)
        return JsonResponse(context)

    return render(request, 'storage/excelimport.html', context)


def select_table(request):
    global wb
    print('wb type: ', type(wb))
    if request.method == 'POST':
        sheet_name = request.POST.get('sheet')
        table_name = request.POST.get('table')
        selected_sheet = wb[sheet_name]
        lookup_table = selected_sheet.tables[table_name]
        data = selected_sheet[lookup_table.ref]
        rows_list = []

        for row in data:
            cols = []
            # print(type(row), '\n')
            for col in row:
                cols.append(col.value)
            rows_list.append(cols)

        df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
        df.to_csv(f'{sheet_name}-{table_name}.csv', index=False)

        data_rows = df.to_dict(orient='records')
        column_names = df.columns.tolist()
        context = {'column_names': column_names, 'data_rows': data_rows}
        return render(request, 'storage/table_display.html', context)
    else:
        return HttpResponse("Method not allowed", status=405)


def sanitize_for_import(value):
    if isinstance(value, str):
        escaped_value = value.replace("'", "''")
        return escaped_value
    elif isinstance(value, list):
        value_list = json.dumps(value)
        return value_list
    elif isinstance(value, dict):
        str_value = str(value)
        escaped_value = str_value.replace("'", '"')
        return escaped_value
    elif isinstance(value, set):
        str_value = str(value)
        escaped_value = str_value.replace("'", "''")
        return escaped_value
    elif not value:
        escaped_value = None
        return escaped_value
    else:
        return value

def load2db(self, df):

    def __sanitize_for_sql(value):
        if isinstance(value, str):
            escaped_value = value.replace("'", "''")
            return escaped_value
        elif isinstance(value, list):
            value_list = json.dumps(value)
            return value_list
        elif isinstance(value, dict):
            str_value = str(value)
            escaped_value = str_value.replace("'", '"')
            return escaped_value
        elif isinstance(value, set):
            str_value = str(value)
            escaped_value = str_value.replace("'", "''")
            return escaped_value
        else:
            return value

    # Establish a connection to the PostgreSQL database:
    dbname = "dg_bae"
    user = "postgres"
    password = "postgres"
    host = "localhost"
    port = "5432"

    connection = psycopg2.connect(
        dbname=dbname,
        user=user,
        password=password,
        host=host,
        port=port
    )

    connection.autocommit = True
    cursor = connection.cursor()

    try:
        # Populate rows
        for _, row in df.iterrows():
            query = f"""INSERT INTO storage_field (
                    field_list_id, 
                    source_list_id, 
                    field_alias, 
                    field_source_type, 
                    field_source_id,
                    field_name, 
                    field_value, 
                    field_function, 
                    function_field_list, 
                    field_description) 
                    VALUES (
                        (select id 
                        from storage_fieldlist 
                        where field_list_name like '{row['field_list']}'),
                    
                        COALESCE((select id 
                        from storage_sourcelist 
                        where source_list_name like '{row['source_list']}'), NULL),
                    
                        '{row['field_alias']}', 
                        '{row['field_source_type']}', 
                    
                        COALESCE((select id
                        from storage_source 
                        where source_alias like '{row['field_source']}' and source_union_list_name_id = (select id 
                        from storage_sourcelist 
                        where source_list_name like '{row['source_list']}')), NULL), 
                    
                        '{row['field_name']}', 
                        '{row['field_value']}', 
                        '{__sanitize_for_sql(row['field_function'])}', 
                        '{row['function_field_list']}', 
                        '{row['field_description']}'
                    );"""

            cursor.execute(query)
            # connection.commit()
        print("Data inserted successfully")
    except Exception as e:
        print(f"Error inserting data: {e}")

        cursor.close()
        connection.close()


@csrf_exempt
def import_excel(request):
    body_unicode = request.body.decode('utf-8')
    body_data = json.loads(body_unicode)

    sheet = body_data.get('sheet')
    table = body_data.get('table')
    file_path = body_data.get('file_path')

    # print("Request data: ", request.body)
    # print("Request file_url: ", request.body['file_url'])
    df = import_table_from_excel(file_path, sheet, table)
    df_dict = df.to_dict(orient='records')
    context = {'text': 'CSV Loaded', 'csv': df_dict}
    return JsonResponse(context)
    # return render(request, 'storage/import_csv.html', context)
    # return redirect('import_csv.html', context)


@csrf_exempt
def import_csv(request):
    body_unicode = request.body.decode('utf-8')
    body_data = json.loads(body_unicode)

    sheet = body_data.get('sheet')
    table = body_data.get('table')
    file_path = body_data.get('file_path')

    df = import_table_from_excel(file_path, sheet, table)

    import_result = []
    if table[:3] == 'FL_':
        for index, row in df.iterrows():
            try:
                source_list, created = SourceList.objects.get_or_create(source_list=row['source_list'])
                if created:
                    import_result.append((row['source_list'], 'SourceList created'))
                    print(f"Field {source_list.source_list} SourceList created")

                field_list, created = FieldList.objects.get_or_create(field_list_name=row['field_list'], data_source=source_list)
                if created:
                    import_result.append((row['field_list'], 'FieldList created'))
                    print(f"Field {field_list.field_list_name} FieldList created")

                if row['field_source']:
                    field_source = Source.objects.get(source_union_list=source_list, source_alias=row['field_source'])
                else:
                    field_source = None
                # sanitized_field_name = sanitize_for_import(row['field_name'])
                # sanitized_field_value = sanitize_for_import(row['field_value'])
                # sanitized_field_function = sanitize_for_import(row['field_function'])
                # sanitized_function_field_list = sanitize_for_import(row['function_field_list'])
                # sanitized_field_description = sanitize_for_import(row['field_description'])

                field, created = Field.objects.update_or_create(
                    field_list=field_list,
                    source_list=source_list,
                    field_alias=row['field_alias'],
                    # field_source=field_source,
                    defaults={
                        'field_list': field_list,
                        'source_list': source_list,
                        'field_alias': row['field_alias'],
                        'field_source_type': row['field_source_type'],
                        'field_source': field_source,
                        'field_name': row['field_name'],
                        'field_value': row['field_value'],
                        'field_function': row['field_function'],
                        'function_field_list': row['function_field_list'],
                        'field_description': row['field_description']
                    }
                )
                if created:
                    import_result.append((row['field_alias'], 'created'))
                    print(f"Field {field.field_alias} created")
                else:
                    import_result.append((row['field_alias'], 'updated'))
                    print(f"Field {field.field_alias} updated")
            except Exception as e:
                import_result.append(f"{row['field_alias']}: {e}")
                print(f"Error inserting data: {e}")


    elif table[:3] == 'DS_':
        for index, row in df.iterrows():
            try:
                source_union_list, created = SourceList.objects.get_or_create(
                    source_list=row['source_union_list_name']
                )
                if created:
                    import_result.append((row['source_union_list_name'], 'SourceList created'))
                    print(f"SourceList {source_union_list.source_list} created")

                if row['source_type'] == 'query':
                    table_name = None
                    source_list = None
                    source_system = None
                    source_scheme = None
                    union_type = UnionType.objects.get(union_type=row['union_type'])
                    query_name, created = Query.objects.get_or_create(
                        query_name=row['source_name']
                    )
                    if created:
                        import_result.append((row['source_name'], 'Query created'))
                        print(f"Query {query_name.query_name} created")
                    else:
                        import_result.append((row['source_name'], 'Query already exist'))
                        print(f"Query {query_name.query_name} already exist")

                elif row['source_type'] == 'data_source':
                    source_list = SourceList.objects.get(source_list=row['source_name'])
                    query_name = None
                    table_name = None
                    source_system = None
                    source_scheme = None
                    union_type = UnionType.objects.get(union_type=row['union_type'])

                elif row['source_type'] == 'table':
                    table_name = row['source_name']
                    source_system = SourceSystem.objects.get(source_system_name=row['source_system'])
                    try:
                        source_scheme = SourceScheme.objects.get(source_scheme_name=row['source_scheme'])
                    except SourceScheme.DoesNotExist:
                        source_scheme = None
                    try:
                        union_type = UnionType.objects.get(union_type=row['union_type'])
                    except UnionType.DoesNotExist:
                        union_type = None
                    query_name = None
                    source_list = None

                source, created = Source.objects.update_or_create(
                    source_union_list=source_union_list,
                    source_alias=row['source_alias'],
                    source_type=row['source_type'],
                    # query_name=query_name,
                    # source_list=source_list,
                    # table_name=table_name,
                    # source_system=row['source_system'],
                    # source_scheme=row['source_scheme'],
                    # union_type=row['union_type'],
                    # union_condition=row['union_condition'],
                    # source_description=row['source_description'],

                    defaults={
                        # 'source_union_list': source_union_list,
                        # 'source_alias': row['source_alias'],
                        # 'source_type': row['source_type'],
                        'query_name': query_name,
                        'source_list': source_list,
                        'table_name': table_name,
                        'source_system': source_system,
                        'source_scheme': source_scheme,
                        'union_type': union_type,
                        'union_condition': row['union_condition'],
                        'source_description': row['source_description']
                    }
                )
                if created:
                    import_result.append((row['source_alias'], 'DataSource created'))
                    print(f"DataSource {source.source_alias} created")
                else:
                    import_result.append((row['source_alias'], 'DataSource updated'))
                    print(f"DataSource {source.source_alias} updated")

            except Exception as e:
                import_result.append(f"{row['source_alias']}: {e}")
                print(f"Error inserting data: {e}")

    elif table[:2] == 'Q_':
        for index, row in df.iterrows():
            query_alias = row.get('query_alias', None)
            try:
                field_list = FieldList.objects.get(field_list_name=row['query_fields'])
            except FieldList.DoesNotExist:
                field_list = None
            try:
                source_list = SourceList.objects.get(source_list=row['query_source'])
            except SourceList.DoesNotExist:
                source_list = None

            try:
                query_name, created = Query.objects.update_or_create(
                    query_name=row['query_name'],
                    defaults={
                        'field_list': field_list,
                        'source_list': source_list,
                        'query_conditions': row['query_conditions'],
                        'query_alias': query_alias,
                        'query_description': row['query_description']
                    }
                )
                if created:
                    import_result.append((row['query_name'], 'created'))
                    print(f"Query {query_name.query_name} created")
                else:
                    import_result.append((row['query_name'], 'updated'))
                    print(f"Query {query_name.query_name} updated")
            except Exception as e:
                import_result.append(f"{row['query_name']}: {e}")
                print(f"Error inserting data: {e}")

    else:
        print(f"This type of object do not support for import, {table}")
        import_result = f"This type of object do not support for import, {table}"

    context = {'import_result': import_result}
    return JsonResponse(context)


# def import_table_from_excel(self, workbook_filename, sheet_name: str = '', table_name: str = ''):
def import_table_from_excel(workbook_filename, sheet_name: str = '', table_name: str = ''):
    wb = load_workbook(filename=workbook_filename)
    print(wb.sheetnames)
    sheet = wb[sheet_name]
    print(sheet.tables.keys(), '\n')
    lookup_table = sheet.tables[table_name]
    print(lookup_table.ref)

    data = sheet[lookup_table.ref]
    rows_list = []

    for row in data:
        cols = []
        for col in row:
            cols.append(col.value)
        rows_list.append(cols)

    df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
    df.to_csv(f'{table_name}.csv', index=False)

    # self.display_df(df)
    return df
